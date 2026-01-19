import pyodbc
import json
import io
import pandas as pd
from google.cloud import bigquery, storage
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import yaml

# ---------------- Config ----------------
def load_yaml_config(gcs_path):
    client = storage.Client()
    bucket_name, blob_name = gcs_path.replace("gs://", "").split("/", 1)
    data = client.bucket(bucket_name).blob(blob_name).download_as_text()
    return yaml.safe_load(data)

def load_config_from_csv(config_gcs_path):
    client = storage.Client()
    bucket_name, blob_name = config_gcs_path.replace("gs://", "").split("/", 1)
    data = client.bucket(bucket_name).blob(blob_name).download_as_text()
    return pd.read_csv(io.StringIO(data))

# ---------------- Secrets ----------------
def get_sqlserver_creds(project_config, v_secret_name):
    secret_path = project_config['env']['v_pwd_secrets_url'] + v_secret_name
    creds = json.loads(access_secret(secret_path))   # assumes access_secret() is defined
    return creds["user"], creds["password"]

# ---------------- SQL Server ----------------
def get_sqlserver_count_checksum(server, database, table, username, passwd):
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={server};DATABASE={database};UID={username};PWD={passwd}"
    )
    query = f"SELECT COUNT(*) AS row_count, CHECKSUM_AGG(BINARY_CHECKSUM(*)) AS checksum FROM {table}"
    with pyodbc.connect(conn_str) as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        row = cursor.fetchone()
        return {"row_count": row[0], "checksum": row[1]}

# ---------------- BigQuery ----------------
BQ_PROJECT = "hca-hin-dev-cur-parallon"
BQ_DATASET = "edwpsc"

def get_bigquery_count_checksum(client, table_name):
    table_full = f"{BQ_PROJECT}.{BQ_DATASET}.{table_name}"
    count_query = f"SELECT COUNT(*) as row_count FROM `{table_full}`"
    row_count = list(client.query(count_query).result())[0].row_count

    checksum_query = f"""
    SELECT TO_HEX(FARM_FINGERPRINT(STRING_AGG(TO_JSON_STRING(t)))) as checksum
    FROM `{table_full}` t
    """
    checksum = list(client.query(checksum_query).result())[0].checksum
    return {"row_count": row_count, "checksum": checksum}

# ---------------- Report ----------------
def write_excel_report(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Server", "Database", "Source_Table", "Target_Table",
        "RowCount_SQL", "RowCount_BQ", "Checksum_SQL", "Checksum_BQ", "Status"
    ])

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for res in results:
        status = "PASS" if (
            res["sql_row_count"] == res["bq_row_count"]
            and str(res["sql_checksum"]) == str(res["bq_checksum"])
        ) else "FAIL"

        ws.append([
            res["server"],
            res["database"],
            res["source_table"],
            res["target_table"],
            res["sql_row_count"],
            res["bq_row_count"],
            res["sql_checksum"],
            res["bq_checksum"],
            status,
        ])

        for cell in ws[ws.max_row]:
            cell.fill = green if status == "PASS" else red

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    passed = sum(
        1 for r in results
        if r["sql_row_count"] == r["bq_row_count"] and str(r["sql_checksum"]) == str(r["bq_checksum"])
    )
    failed = total - passed
    ws2.append(["Total Tables", "Passed", "Failed"])
    ws2.append([total, passed, failed])

    wb.save(output_path)

# ---------------- Upload ----------------
def upload_to_gcs(bucket_name, file_path, blob_name):
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    bucket.blob(blob_name).upload_from_filename(file_path)

# ---------------- Runner ----------------
def run_validation(config_csv, project_config_path, gcs_output_bucket, gcs_output_file, v_secret_name, batch_size=50):
    project_config = load_yaml_config(project_config_path)
    config_df = load_config_from_csv(config_csv)
    bq_client = bigquery.Client()

    username, passwd = get_sqlserver_creds(project_config, v_secret_name)

    results = []
    for start in range(0, len(config_df), batch_size):
        batch_df = config_df.iloc[start:start + batch_size]

        for _, row in batch_df.iterrows():
            if not row["validate"]:
                continue

            print(f"Validating {row['server']}.{row['database']}.{row['source_table']} vs {row['target_table']}")

            sql_stats = get_sqlserver_count_checksum(
                row["server"], row["database"], row["source_table"], username, passwd
            )
            bq_stats = get_bigquery_count_checksum(bq_client, row["target_table"])

            results.append({
                "server": row["server"],
                "database": row["database"],
                "source_table": row["source_table"],
                "target_table": row["target_table"],
                "sql_row_count": sql_stats["row_count"],
                "sql_checksum": sql_stats["checksum"],
                "bq_row_count": bq_stats["row_count"],
                "bq_checksum": bq_stats["checksum"],
            })

    local_report = "/tmp/final_count_checksum_report.xlsx"
    write_excel_report(results, local_report)
    upload_to_gcs(gcs_output_bucket, local_report, gcs_output_file)

    print(f"âœ… Report generated: gs://{gcs_output_bucket}/{gcs_output_file}")